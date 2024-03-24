#!/usr/bin/python
#
from openpyxl import load_workbook
from os import listdir,path,rename
import re

def number_c2e(chinese_number):
  num_map = dict(一=1,二=2,三=3,四=4,五=5,六=6,七=7,八=8,九=9,十=10)
  size = len(chinese_number)
  if size == 0: return 0
  if size < 2:
      return num_map[chinese_number]
  ans = 0
  continue_flag = False
  for i in range(size):
    if continue_flag:
      continue_flag = False
      continue
    if i + 1 < size and chinese_number[i + 1] == '十':
      ans += num_map[chinses_number[i]] * 10
      continue_flag = True
      continue
    ans += num_map[chinese_number[i]]
  return ans

def cn_num2alb_num(filename):
  di = re.search('第',filename).start()
  ye = re.search('页',filename).start()
  if di + 1 == ye - 1:
    cn_num = filename[di + 1] 
    return number_c2e(cn_num)
  else:
    first_cn = filename[di + 1]
    second_cn = filename[ye - 1]
    return number_c2e(first_cn+second_cn)

xlsxName = '绘本批量导入模板.xlsx'
wb = load_workbook(xlsxName)
sheet = wb.active
max_row = sheet.max_row + 1

worksDir = []
for i in listdir('./'):
  if path.isdir(i):
    worksDir.append(i)
print('当前目录包含的作品：')
print(worksDir)
issort = input('是否进行排序(y/n)：')
if issort == 'y':
  worksDir.sort(key = lambda l: int(re.findall('\d+',l)[0]))
#worksDir.sort()

for i in range(2,max_row):
  print('当前目录包含的作品：')
  print(worksDir)
  stuNum = i - 1
  stuName = sheet['a' + str(i)].value
  workName = sheet['h' + str(i)].value

  for dir in worksDir:
    #print(dir)
    if dir.startswith(str(stuNum)):
      curDir = dir
      break
    elif dir.startswith(stuName):
      curDir = dir
      break
  #print(curDir)
  print('正在操作的目录：' + curDir + ' stuNum: ' + str(stuNum) + ' stuName: ' + stuName + ' workName: ' + workName)

  worksFile = []
  for w in listdir(curDir):
    if path.isfile(curDir + '/' + w):
      #print(w)
      if w.lower().endswith(".jpg") or w.lower().endswith(".png") or w.lower().endswith("jpeg"):
        if '封面' in w:
          firstPic = w
        elif '封底' in w or '背面' in w:
          lastPic = w
        elif '扉页' in w:
          secondPic = w
        elif '长篇' in w:
          continue
        else:
          worksFile.append(w)
  #print(worksFile)

  for l in worksFile:
    print(l)
  print()
  #sortIdx = int(input('输入排序索引下标：')) 
  #if sortIdx == 0:
  #worksFile.sort(key = lambda p: int(re.findall('\d+',p)[sortIdx]))
  worksFile.sort(key = lambda p: int(cn_num2alb_num(p)))
  #worksFile.sort()
  if 'firstPic' in locals():
    worksFile.insert(0,firstPic)
  if 'lastPic' in locals():
    worksFile.append(lastPic)
  if 'secondPic' in locals():
    worksFile.insert(1,secondPic)
  for l in worksFile:
    print(l)
  print('以上将生成' + str(len(worksFile)) + '条')
  gen = int(input('是否使用上面列表进行生成(是：1，否：2): '))
  if gen == 1:
    fileCount = len(worksFile) + 1
  else:
    fileCount = int(input('输入一个文件个数: ')) + 1

  #fileCount = len(worksFile) + 1
  workFileNameList = []
  for j in range(1,fileCount):
    workFileNameList.append(str(stuNum) + str(stuName) + '《' + workName + '》'+ str(j) + '.jpg')

  for y in range(0,len(workFileNameList)):
    oldName = curDir + '/' + worksFile[y]
    newName = curDir + '/' + workFileNameList[y]
    rename(oldName,newName)
    print('将 ' + oldName + ' 重命名成 ' + newName)

  fileNameList = ''
  for x in workFileNameList:
    fileNameList = fileNameList + x + ','
  sheet['j' + str(i)] = fileNameList
  wb.save(xlsxName)
  print(workFileNameList)
  print(fileNameList)
  print()
  if 'firstPic' in locals():
    del(firstPic)
  if 'lastPic' in locals():
    del(lastPic)
  if 'secondPic' in locals():
    del(secondPic)
  del worksFile[:]
  del workFileNameList[:]

